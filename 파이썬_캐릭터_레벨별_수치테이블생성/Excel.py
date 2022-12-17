from openpyxl import load_workbook
from openpyxl import Workbook
import plotly.express as px


x_data = []
y_data = [100, 150, 200, 250, 300]
agentName = []
data = []
level_data = []
countColumn = 0
level = 1
agentCount = 6
temp = 1;
tempInt = 0
a = 0
load_wb = load_workbook("전투시스템기획서_밸런싱시트_B977082백정열.xlsx", data_only=True)

load_ws = load_wb['요원별속성테이블']

write_wb = Workbook()
write_ws = write_wb.create_sheet('생성시트')
write_ws = write_wb.active

while level<26:
    for i in range(agentCount):
        for j in range(3):
            a = j
            a = a + 2 + j
            tData = load_ws.cell(row=i + 2, column=a)
            data.append(tData)
    for i in range(agentCount):
        for j in range(3):
            a = j
            a = a + 3 + j
            tData = load_ws.cell(row=i + 2, column=a).value
            tData = data[tempInt].value + int(tData) * (level - 1)
            tempInt += 1
            level_data.append(tData)
    write_ws.cell(row=1, column=2+((level-1)*3), value=str(level)+'레벨 체력')
    write_ws.cell(row=1, column=3+((level-1)*3), value=str(level)+'레벨 총기숙련도')
    write_ws.cell(row=1, column=4+((level-1)*3), value=str(level)+'레벨 속도')
    a = (level-1) * 18
    for i in range(agentCount):
        for j in range (int(int(len(level_data))/level/ agentCount)):
            write_ws.cell(row=i + 2, column=j + 2+(level-1)*3, value=level_data[a])
            a = a + 1
    level = level+1

write_ws.cell(row=1, column=1, value='이름')
write_ws.cell(row=2, column=1, value='기본요원A')
write_ws.cell(row=3, column=1, value='기본요원B')
write_ws.cell(row=4, column=1, value='기본요원C')
write_ws.cell(row=5, column=1, value='기본요원D')
write_ws.cell(row=6, column=1, value='기본요원E')
write_ws.cell(row=7, column=1, value='기본요원F')
write_wb.save('output.xlsx')

#fig=px.scatter(x=x_data, y=y_data)
#fig.show()
